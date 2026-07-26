#!/usr/bin/env python3
"""Build a compact CONUS power-plant layer from EIA-860M.

The monthly EIA-860M workbook reports one row per generator.  This builder
combines the ``Operating`` and ``Planned`` worksheets into one GeoJSON point
per EIA plant while preserving separate unit counts, status counts and
nameplate capacities.  It deliberately excludes Alaska, Hawaii and Puerto
Rico because the atlas region is the continental United States.

The input workbook is not redistributed.  The generated GeoJSON and metadata
record its SHA-256 digest, release date, source URL and transformation rules.
``openpyxl`` is required only to run this offline builder.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import shutil
import subprocess
import tempfile
import urllib.request
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as error:  # pragma: no cover - dependency error is explicit
    raise SystemExit(
        "openpyxl is required to read EIA-860M. Install it in the builder "
        "environment and rerun this script."
    ) from error


SOURCE_ID = "us-eia860m-2026-06"
SOURCE_PAGE_URL = "https://www.eia.gov/electricity/data/eia860m/"
SOURCE_DOWNLOAD_URL = (
    "https://www.eia.gov/electricity/data/eia860m/xls/"
    "june_generator2026.xlsx"
)
SOURCE_PERIOD = "2026-06"
RELEASE_DATE = "2026-07-23"
SOURCE_LICENSE = "U.S. Government public domain"
SOURCE_LICENSE_URL = "https://www.eia.gov/about/copyrights_reuse.php"

# The atlas region is explicitly the continental United States.  These bounds
# match its region button and remove non-CONUS points before aggregation.
CONUS_BOUNDS = {
    "west": -125.0,
    "south": 24.0,
    "east": -66.0,
    "north": 50.0,
}
EXCLUDED_STATES = {"AK", "HI", "PR"}
WORKSHEETS = ("Operating", "Planned")


def parse_args() -> argparse.Namespace:
    project_root = Path(__file__).resolve().parents[2]
    output = project_root / "data" / "grid-atlas" / "us-eia860m-plants.geojson"
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-xlsx",
        type=Path,
        help="Use an already-downloaded EIA-860M workbook.",
    )
    parser.add_argument(
        "--url",
        default=SOURCE_DOWNLOAD_URL,
        help="Official EIA workbook URL.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=output,
        help=f"GeoJSON destination (default: {output}).",
    )
    parser.add_argument(
        "--metadata-output",
        type=Path,
        help="Metadata destination; defaults beside --output.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=180.0,
        help="Download timeout in seconds.",
    )
    return parser.parse_args()


def sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def download_with_urllib(url: str, timeout: float) -> Path:
    request = urllib.request.Request(
        url,
        headers={
            "Accept": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet, application/octet-stream;q=0.9"
            ),
            "User-Agent": (
                "Comunicacion-Grid-Atlas/1.0 "
                "(+https://qdgvr.github.io/fecundidad.github.io/)"
            ),
        },
    )
    temporary = tempfile.NamedTemporaryFile(
        prefix="eia860m-", suffix=".xlsx", delete=False
    )
    temporary_path = Path(temporary.name)
    try:
        with temporary, urllib.request.urlopen(request, timeout=timeout) as response:
            while chunk := response.read(1024 * 1024):
                temporary.write(chunk)
        return temporary_path
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise


def download_with_curl(url: str, timeout: float) -> Path:
    curl = shutil.which("curl")
    if not curl:
        raise RuntimeError("curl is not available")
    temporary = tempfile.NamedTemporaryFile(
        prefix="eia860m-", suffix=".xlsx", delete=False
    )
    temporary_path = Path(temporary.name)
    temporary.close()
    try:
        subprocess.run(
            [
                curl,
                "--fail",
                "--location",
                "--silent",
                "--show-error",
                "--max-time",
                str(max(1, math.ceil(timeout))),
                "--output",
                str(temporary_path),
                url,
            ],
            check=True,
        )
        return temporary_path
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise


def acquire_workbook(args: argparse.Namespace) -> tuple[Path, bool, str]:
    if args.input_xlsx:
        path = args.input_xlsx.expanduser().resolve()
        if not path.is_file():
            raise SystemExit(f"Input workbook does not exist: {path}")
        return path, False, "local input"

    try:
        return download_with_urllib(args.url, args.timeout), True, "urllib"
    except Exception as urllib_error:
        try:
            return download_with_curl(args.url, args.timeout), True, "curl"
        except Exception as curl_error:
            raise SystemExit(
                "Could not download EIA-860M with urllib or curl: "
                f"{urllib_error}; {curl_error}"
            ) from curl_error


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def finite_number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def integer(value: Any) -> int | None:
    number = finite_number(value)
    if number is None or not number.is_integer():
        return None
    return int(number)


def status_code(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    match = re.match(r"^\(([A-Z]+)\)", text)
    return match.group(1) if match else text


def within_conus(longitude: float, latitude: float, state: str | None) -> bool:
    if state in EXCLUDED_STATES:
        return False
    return (
        CONUS_BOUNDS["west"] <= longitude <= CONUS_BOUNDS["east"]
        and CONUS_BOUNDS["south"] <= latitude <= CONUS_BOUNDS["north"]
    )


def first_nonempty(current: str | None, candidate: Any) -> str | None:
    return current or clean_text(candidate)


def empty_plant(plant_id: int) -> dict[str, Any]:
    return {
        "plant_id": plant_id,
        "name": None,
        "state": None,
        "county": None,
        "entity_id": None,
        "entity_name": None,
        "balancing_authority": None,
        "sector": None,
        "coordinates": Counter(),
        "technologies": set(),
        "energy_sources": set(),
        "operating_statuses": Counter(),
        "planned_statuses": Counter(),
        "current_units": 0,
        "planned_units": 0,
        "current_nameplate_mw": 0.0,
        "current_summer_mw": 0.0,
        "available_nameplate_mw": 0.0,
        "planned_nameplate_mw": 0.0,
        "planned_summer_mw": 0.0,
        "earliest_operating_year": None,
        "earliest_planned_year": None,
    }


def update_minimum(current: int | None, candidate: Any) -> int | None:
    value = integer(candidate)
    if value is None:
        return current
    return value if current is None else min(current, value)


def iter_sheet_rows(
    worksheet: Any,
) -> tuple[dict[str, int], Iterable[tuple[Any, ...]]]:
    rows = worksheet.iter_rows(values_only=True)
    next(rows, None)
    next(rows, None)
    header_row = next(rows, None)
    if not header_row:
        raise ValueError(f"Worksheet {worksheet.title} has no header row")
    headers = {
        clean_text(value): index
        for index, value in enumerate(header_row)
        if clean_text(value)
    }
    required = {
        "Plant ID",
        "Plant Name",
        "Plant State",
        "Latitude",
        "Longitude",
        "Nameplate Capacity (MW)",
        "Technology",
        "Energy Source Code",
        "Status",
    }
    missing = sorted(required - headers.keys())
    if missing:
        raise ValueError(
            f"Worksheet {worksheet.title} is missing columns: {', '.join(missing)}"
        )
    return headers, rows


def cell(row: tuple[Any, ...], headers: dict[str, int], name: str) -> Any:
    index = headers.get(name)
    return row[index] if index is not None and index < len(row) else None


def aggregate_workbook(path: Path) -> tuple[dict[int, dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    plants: dict[int, dict[str, Any]] = {}
    stats: dict[str, Any] = {
        "source_rows": {},
        "valid_conus_rows": {},
        "invalid_coordinate_rows": {},
        "outside_region_rows": {},
    }

    try:
        for sheet_name in WORKSHEETS:
            worksheet = workbook[sheet_name]
            headers, rows = iter_sheet_rows(worksheet)
            stats["source_rows"][sheet_name] = 0
            stats["valid_conus_rows"][sheet_name] = 0
            stats["invalid_coordinate_rows"][sheet_name] = 0
            stats["outside_region_rows"][sheet_name] = 0

            for row in rows:
                if not any(value is not None for value in row):
                    continue
                stats["source_rows"][sheet_name] += 1
                plant_id = integer(cell(row, headers, "Plant ID"))
                latitude = finite_number(cell(row, headers, "Latitude"))
                longitude = finite_number(cell(row, headers, "Longitude"))
                state = clean_text(cell(row, headers, "Plant State"))
                if plant_id is None or latitude is None or longitude is None:
                    stats["invalid_coordinate_rows"][sheet_name] += 1
                    continue
                if not within_conus(longitude, latitude, state):
                    stats["outside_region_rows"][sheet_name] += 1
                    continue

                stats["valid_conus_rows"][sheet_name] += 1
                plant = plants.setdefault(plant_id, empty_plant(plant_id))
                plant["name"] = first_nonempty(
                    plant["name"], cell(row, headers, "Plant Name")
                )
                plant["state"] = first_nonempty(plant["state"], state)
                plant["county"] = first_nonempty(
                    plant["county"], cell(row, headers, "County")
                )
                plant["entity_id"] = plant["entity_id"] or integer(
                    cell(row, headers, "Entity ID")
                )
                plant["entity_name"] = first_nonempty(
                    plant["entity_name"], cell(row, headers, "Entity Name")
                )
                plant["balancing_authority"] = first_nonempty(
                    plant["balancing_authority"],
                    cell(row, headers, "Balancing Authority Code"),
                )
                plant["sector"] = first_nonempty(
                    plant["sector"], cell(row, headers, "Sector")
                )
                plant["coordinates"][(round(longitude, 6), round(latitude, 6))] += 1

                technology = clean_text(cell(row, headers, "Technology"))
                energy_source = clean_text(cell(row, headers, "Energy Source Code"))
                if technology:
                    plant["technologies"].add(technology)
                if energy_source:
                    plant["energy_sources"].add(energy_source)

                nameplate = finite_number(
                    cell(row, headers, "Nameplate Capacity (MW)")
                ) or 0.0
                summer = finite_number(
                    cell(row, headers, "Net Summer Capacity (MW)")
                ) or 0.0
                code = status_code(cell(row, headers, "Status")) or "UNKNOWN"
                if sheet_name == "Operating":
                    plant["current_units"] += 1
                    plant["current_nameplate_mw"] += nameplate
                    plant["current_summer_mw"] += summer
                    plant["operating_statuses"][code] += 1
                    if code != "OS":
                        plant["available_nameplate_mw"] += nameplate
                    plant["earliest_operating_year"] = update_minimum(
                        plant["earliest_operating_year"],
                        cell(row, headers, "Operating Year"),
                    )
                else:
                    plant["planned_units"] += 1
                    plant["planned_nameplate_mw"] += nameplate
                    plant["planned_summer_mw"] += summer
                    plant["planned_statuses"][code] += 1
                    plant["earliest_planned_year"] = update_minimum(
                        plant["earliest_planned_year"],
                        cell(row, headers, "Planned Operation Year"),
                    )
    finally:
        workbook.close()

    return plants, stats


def rounded_capacity(value: float) -> float:
    return round(value, 3)


def encode_counter(counter: Counter[str]) -> str | None:
    if not counter:
        return None
    return "|".join(f"{key}:{counter[key]}" for key in sorted(counter))


def feature_for_plant(plant: dict[str, Any]) -> dict[str, Any]:
    coordinates = plant["coordinates"].most_common()
    longitude, latitude = coordinates[0][0]
    coordinate_variants = len(coordinates)
    current_mw = rounded_capacity(plant["current_nameplate_mw"])
    planned_mw = rounded_capacity(plant["planned_nameplate_mw"])
    display_mw = max(current_mw, planned_mw)
    # Repeated GeoJSON property names dominate file size at 15k+ points.  The
    # short keys below are lossless and are documented in the metadata
    # sidecar.  Provenance remains on every feature rather than only at
    # collection level.
    properties = {
        "src": SOURCE_ID,
        "i": plant["plant_id"],
        "n": plant["name"],
        "st": plant["state"],
        "co": plant["county"],
        "ei": plant["entity_id"],
        "en": plant["entity_name"],
        "ba": plant["balancing_authority"],
        "se": plant["sector"],
        "cu": plant["current_units"],
        "pu": plant["planned_units"],
        "cm": current_mw,
        "cs": rounded_capacity(plant["current_summer_mw"]),
        "am": rounded_capacity(plant["available_nameplate_mw"]),
        "pm": planned_mw,
        "psm": rounded_capacity(plant["planned_summer_mw"]),
        "dm": display_mw,
        "os": encode_counter(plant["operating_statuses"]),
        "ps": encode_counter(plant["planned_statuses"]),
        "t": "|".join(sorted(plant["technologies"])) or None,
        "f": "|".join(sorted(plant["energy_sources"])) or None,
        "oy": plant["earliest_operating_year"],
        "py": plant["earliest_planned_year"],
        "cv": coordinate_variants,
        "sp": SOURCE_PERIOD,
        "rd": RELEASE_DATE,
        "url": SOURCE_PAGE_URL,
        "lic": SOURCE_LICENSE,
        "ev": "reported-preliminary",
        "gc": (
            "reported-plant-coordinate"
            if coordinate_variants == 1
            else "reported-most-common-plant-coordinate"
        ),
    }
    return {
        "type": "Feature",
        "id": f"eia860m:{plant['plant_id']}",
        "geometry": {
            "type": "Point",
            "coordinates": [longitude, latitude],
        },
        "properties": {key: value for key, value in properties.items() if value is not None},
    }


def portable_path(path: Path) -> str:
    project_root = Path(__file__).resolve().parents[2]
    try:
        return path.resolve().relative_to(project_root).as_posix()
    except ValueError:
        return path.name


def write_json(path: Path, data: Any, compact: bool) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as destination:
        if compact:
            json.dump(
                data,
                destination,
                ensure_ascii=False,
                separators=(",", ":"),
            )
        else:
            json.dump(data, destination, ensure_ascii=False, indent=2)
            destination.write("\n")


def main() -> int:
    args = parse_args()
    workbook_path, temporary, transport = acquire_workbook(args)
    output_path = args.output.expanduser().resolve()
    metadata_path = (
        args.metadata_output.expanduser().resolve()
        if args.metadata_output
        else output_path.with_name("us-eia860m-plants.metadata.json")
    )

    try:
        source_sha256 = sha256_path(workbook_path)
        plants, stats = aggregate_workbook(workbook_path)
        features = [
            feature_for_plant(plants[plant_id])
            for plant_id in sorted(plants)
        ]
        collection = {
            "type": "FeatureCollection",
            "name": "EIA-860M June 2026 continental U.S. power plants",
            "source": {
                "id": SOURCE_ID,
                "url": SOURCE_PAGE_URL,
                "download_url": args.url,
                "period": SOURCE_PERIOD,
                "release_date": RELEASE_DATE,
                "license": SOURCE_LICENSE,
                "license_url": SOURCE_LICENSE_URL,
                "source_sha256": source_sha256,
            },
            "features": features,
        }
        write_json(output_path, collection, compact=True)

        counts = {
            "plant_features": len(features),
            "plants_with_current_units": sum(
                feature["properties"]["cu"] > 0
                for feature in features
            ),
            "plants_with_planned_units": sum(
                feature["properties"]["pu"] > 0
                for feature in features
            ),
            "current_generator_rows": sum(
                feature["properties"]["cu"] for feature in features
            ),
            "planned_generator_rows": sum(
                feature["properties"]["pu"] for feature in features
            ),
            "coordinate_conflict_plants": sum(
                feature["properties"]["cv"] > 1
                for feature in features
            ),
        }
        metadata = {
            "schema_version": 1,
            "built_at": datetime.now(timezone.utc).isoformat(),
            "source": {
                "id": SOURCE_ID,
                "publisher": "U.S. Energy Information Administration",
                "title": "Preliminary Monthly Electric Generator Inventory",
                "period": SOURCE_PERIOD,
                "release_date": RELEASE_DATE,
                "page_url": SOURCE_PAGE_URL,
                "download_url": args.url,
                "source_sha256": source_sha256,
                "download_transport": transport,
                "license": SOURCE_LICENSE,
                "license_url": SOURCE_LICENSE_URL,
            },
            "scope": {
                "region": "continental United States",
                "bounds": CONUS_BOUNDS,
                "excluded_states": sorted(EXCLUDED_STATES),
                "worksheets": list(WORKSHEETS),
            },
            "transform": {
                "unit_of_observation": "one GeoJSON point per EIA Plant ID",
                "coordinate_rule": (
                    "most frequent reported longitude/latitude pair when a "
                    "plant has more than one coordinate pair"
                ),
                "capacity_rule": (
                    "sum generator nameplate and summer capacities separately "
                    "for Operating and Planned worksheets"
                ),
                "status_rule": (
                    "preserve EIA leading status codes and generator counts; "
                    "Operating status OS is excluded only from "
                    "AVAILABLE_NAMEPLATE_MW"
                ),
                "telemetry": "none; EIA-860M is a dated inventory",
            },
            "property_schema": {
                "src": "source registry identifier",
                "i": "original EIA Plant ID",
                "n": "plant name",
                "st": "two-letter state code",
                "co": "county",
                "ei": "EIA entity ID",
                "en": "entity name",
                "ba": "balancing authority code",
                "se": "sector",
                "cu": "generator rows in Operating",
                "pu": "generator rows in Planned",
                "cm": "current nameplate capacity, MW",
                "cs": "current net summer capacity, MW",
                "am": "available nameplate capacity excluding status OS, MW",
                "pm": "planned nameplate capacity, MW",
                "psm": "planned net summer capacity, MW",
                "dm": "display capacity: max(cm, pm), MW",
                "os": "Operating status-code counts separated by |",
                "ps": "Planned status-code counts separated by |",
                "t": "technologies separated by |",
                "f": "energy-source codes separated by |",
                "oy": "earliest reported operating year",
                "py": "earliest reported planned operation year",
                "cv": "distinct reported coordinate pairs for the Plant ID",
                "sp": "source reporting period",
                "rd": "source release date",
                "url": "official source page URL",
                "lic": "source reuse label",
                "ev": "evidence class",
                "gc": "geometry confidence",
            },
            "counts": counts,
            "row_audit": stats,
            "output": {
                "path": portable_path(output_path),
                "bytes": output_path.stat().st_size,
                "sha256": sha256_path(output_path),
            },
        }
        write_json(metadata_path, metadata, compact=False)
        print(
            f"Wrote {len(features):,} EIA plant features to "
            f"{portable_path(output_path)}."
        )
        print(
            f"Current rows: {counts['current_generator_rows']:,}; "
            f"planned rows: {counts['planned_generator_rows']:,}; "
            f"coordinate conflicts: {counts['coordinate_conflict_plants']:,}."
        )
    finally:
        if temporary:
            workbook_path.unlink(missing_ok=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
